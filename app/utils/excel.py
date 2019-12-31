# coding=utf-8
# email:  younixiao@qq.com
# create: 2017年07月20日11:27:15
from cStringIO import StringIO
# from io import BytesIO
import csv
import xlwt
import xlsxwriter
import datetime
import chardet
import sys
import shutil
import os
from openpyxl import Workbook
import json
from utils.dlog import logger
import conf
review_dir = os.path.join(conf.EXCEL_PATH, 'review_export')


def csv_export(data, th_name_list, file_name='export.csv') -> dict:
    """
    导出 csv 文件
    :param data: (list) csv内容
    :param th_name_list: (list) csv 表头, 格式为 [{'name': '邮箱', 'key': 'email'},'ip']
    :param file_name: (str) 文件名
    :return: 返回 csv 文件 http 响应
    """
    th_text_list = []
    th_key_list = []
    for th in th_name_list:
        if isinstance(th, (str, unicode)):
            th_text_list.append(th)
            th_key_list.append(th)
        else:
            th_text_list.append(th['name'])
            th_key_list.append(th['key'])
    si = StringIO()
    cw = csv.writer(si)
    cw.writerow(th_text_list)
    for item in data:
        row = []
        for th in th_key_list:
            row.append(item[th])
        cw.writerow(row)
    return {'data': si.getvalue().encode('utf-8'), 'file_name': file_name}


def xls_export(data, th_name_list, file_name='export.xls'):
    th_text_list = []
    th_key_list = []
    for th in th_name_list:
        if isinstance(th, (str, unicode)):
            th_text_list.append(th)
            th_key_list.append(th)
        else:
            th_text_list.append(th['name'])
            th_key_list.append(th['key'])
    result_list = ','.join([str(th) for th in th_text_list])
    for result in data:
        cells = [str(result[key]).replace('\n', '').replace('\r', '').replace(',', '，') for key in th_key_list]
        result_list += '\n' + ','.join(cells)
    data = dict(data=result_list, file_name=file_name)
    return data


def xls_save(data, th_name_list, file_name='export.xls'):
    """ 存储xls到本地路径"""
    workbook = save_to_workbook(data, th_name_list)
    workbook.save(file_name)


def xls_save_to_api(data, th_name_list, file_name='export.xls'):
    """ 存储xls到本地路径并且移动到api 下的review_export"""
    workbook = save_to_workbook(data, th_name_list)
    workbook.save(file_name)
    if not path.exists(review_dir):
        os.makedirs(review_dir)
    shutil.move(file_name, path.join(review_dir, file_name))
    excel_file_path = path.join(review_dir, file_name).replace('/web/api', '')
    return dict(path=excel_file_path, file_name=file_name)


def save_to_workbook(data, th_name_list):
    workbook = xlwt.Workbook(encoding='utf-8')
    if isinstance(data, list):
        add_sheet('汇总信息', data, th_name_list, workbook)
    elif isinstance(data, object):
        for key, value in data.items():
            add_sheet(key, value, th_name_list, workbook)
    return workbook


def add_sheet(sheet_name, sheet_data, table_header, workbook):
    sheet = workbook.add_sheet(sheet_name)
    style = xlwt.XFStyle()
    font = xlwt.Font()
    font.name = 'SimSun'
    style.font = font

    for index, th_name in enumerate(table_header):
        sheet.write(0, index, th_name['name'], style)

    for item_idx, item in enumerate(sheet_data):
        for th_idx, th_name in enumerate(table_header):
            sheet.write(item_idx + 1, th_idx, item.get(th_name['key'], ''), style,)


# 更换新的插件 使用openptxl
def xlsx_save_to_api(data, th_name_list, file_name='export.xlsx'):
    """ 存储xls到本地路径并且移动到api 下的review_export"""
    try:
        workbook = xlsx_save_to_workbook(data, th_name_list)
        sheet = workbook.get_sheet_by_name('Sheet')
        workbook.remove(sheet)  # 去除默认的新增的数据
        workbook.save(file_name)
        if not path.exists(review_dir):
            os.makedirs(review_dir)
        shutil.move(file_name, path.join(review_dir, file_name))
        excel_file_path = path.join(review_dir, file_name).replace('/web/api', '')
        logger.info('excel_file_path')
        logger.info(excel_file_path)
        return dict(path=excel_file_path, file_name=file_name)
    except Exception as e:
        logger.error('导出审核报表失败原因：' + str(e))
        return ('导出审核报表失败', 500)


def xlsx_save_to_workbook(data, th_name_list):
    workbook = Workbook()
    if isinstance(data, list):
        xlsx_add_sheet(value_to_unicode('汇总信息'), data, th_name_list, workbook)
    elif isinstance(data, object):
        for key, value in data.items():
            xlsx_add_sheet(key, value, th_name_list, workbook)
    return workbook


def xlsx_add_sheet(sheet_name, sheet_data, table_header, workbook):
    sheet = workbook.create_sheet(value_to_unicode(sheet_name))
    for index, th_name in enumerate(table_header):
        sheet.cell(row=1, column=index + 1).value = value_to_unicode(th_name['name'])
    for item_idx, item in enumerate(sheet_data):
        for th_idx, th_name in enumerate(table_header):
            sheet.cell(row=item_idx + 2, column=th_idx + 1).value = value_to_unicode(item.get(th_name['key'], ''))


def value_to_unicode(value):
    if value:
        try:
            return unicode(str(value), 'utf8')
        except:
            return value
    else:
        return value


def json2excel(file_name, data, is_json_dumps=False):
    """
    将 json 数据生成 excel
    :param file_name:
    :param data: {header: [{title: '', key: ''}], content: [{[key]: ''}]}
    :return:
    """
    header = data.get('header', [])
    content = data.get('content', [])
    # Create an new Excel file and add a worksheet.
    workbook = xlsxwriter.Workbook(file_name)
    worksheet = workbook.add_worksheet()
    # Add a bold format to use to highlight cells.
    bold = workbook.add_format({'bold': True})
    for header_index, header_item in enumerate(header):
        key = header_item.get('key', '')
        title = header_item.get('title', '')
        # 设置列宽
        worksheet.set_column(0, header_index, 32)
        # 生成列名
        worksheet.write(0, header_index, title, bold)
        # 生成内容
        for content_index, content_item in enumerate(content):
            value = content_item.get(key, '')
            if is_json_dumps:
                worksheet.write(content_index + 1, header_index, json.dumps(value, ensure_ascii=False))
            else:
                worksheet.write(content_index + 1, header_index, str(value))
    workbook.close()
